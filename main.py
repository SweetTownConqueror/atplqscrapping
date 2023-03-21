from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import json
from pathlib import Path
from selenium.webdriver.common.by import By
import openpyxl


URL = ""
EXEL_SHEETNAME='pof'
NUMBER_OF_QUESTION_TO_SCRAP=3
COOKIE_PATH = '../cookie.json'



DRIVER_PATH = '..\\chromedriver.exe'
#driver = webdriver.Chrome(executable_path=DRIVER_PATH)
options = Options()
#hide or not windows
options.headless = False
#options.add_argument("--window-size=1920,1200")
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options, executable_path=DRIVER_PATH)
driver.implicitly_wait(10)
driver.get(URL)


def get_cookie():
    # retrieve cookies from a json file
    for cookie in json.loads(Path(COOKIE_PATH).read_text()):
        if 'sameSite' in cookie:
            print(cookie['sameSite'])
            if cookie['sameSite'] == None:
                cookie['sameSite'] = 'Strict'
            if cookie['sameSite'] == 'no_restriction':
                cookie['sameSite'] = 'None'
            if cookie['sameSite'] == 'lax':
                cookie['sameSite'] = 'Lax'
            if cookie['domain'] == 'mail.google.com':
                cookie['domain'] = 'google.com'
        #cookie["secure"]= False
        #cookie["domain"]=".google.com"
        try:
            driver.add_cookie(cookie)
        except Exception as e:
            print("eerr i",e)
    
def writeToExcelDoc(_raw, _col, _item, _worksheet_name):
    workbook = openpyxl.load_workbook('../last200fr.xlsx')
    sheets = workbook.sheetnames
    worksheet = workbook[sheets[sheets.index(_worksheet_name)]]
    worksheet.cell(row = _raw, column = _col, value = _item)

    workbook.save("../last200fr.xlsx")
    
def scrap_page(_nb_questions, _worksheet_name):
    driver.get(URL)
    raw = 1
    for i in range(_nb_questions):
        nextQuestion = driver.find_element(By.ID, 'next-question')
        questions = driver.find_elements(By.XPATH, "//div[@id='move']/div[@class='questionText']")
        responses = driver.find_elements(By.XPATH, "//div[@id='move']/ul/li")
        goodResponse = ""
        col = 1
        for q in questions:
            question = q.text
            print(question)
            writeToExcelDoc(raw, col, question, _worksheet_name)
            col += 1
        for r in responses:
            response = r.text
            print(response)
            if 'answerColorTrue' in r.get_attribute('class').split():
                goodResponse = r.text[0]
                print('bonne reponse: '+goodResponse)
                writeToExcelDoc(raw, 6, goodResponse, _worksheet_name)
            writeToExcelDoc(raw, col, response, _worksheet_name)
            col += 1
        raw +=1
        nextQuestion.click()

get_cookie()
scrap_page(NUMBER_OF_QUESTION_TO_SCRAP, EXEL_SHEETNAME)
driver.quit()



